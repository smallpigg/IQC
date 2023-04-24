import openpyxl

def concatenate_columns(file_path, sheet_name, col1, col2, delimiter, new_col):
    """
    将Excel工作表中的两列字符合并，用指定字符串连接。

    :param file_path: Excel文件的路径
    :param sheet_name: 要操作的工作表名称
    :param col1: 第一列的索引（基于1的索引）
    :param col2: 第二列的索引（基于1的索引）
    :param delimiter: 用于连接字符串的分隔符
    :param new_col: 合并后的新列的索引（基于1的索引）
    """
    # 加载工作簿和工作表
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # 获取工作表的最大行数
    max_row = sheet.max_row

    for row in range(1, max_row + 1):
        # 获取两列的值
        value1 = sheet.cell(row=row, column=col1).value
        value2 = sheet.cell(row=row, column=col2).value

        # 合并值并添加分隔符
        combined_value = str(value1) + delimiter + str(value2)

        # 将合并后的值写入新列
        sheet.cell(row=row, column=new_col).value = combined_value

    # 保存更改后的工作簿
    workbook.save(file_path)
    print(f"合并了{max_row}行")

# 示例用法
# file_path = "your_excel_file.xlsx"  # 你的Excel文件路径
# sheet_name = "Sheet1"  # 你要操作的工作表名称
# col1 = 1  # 第一列的索引（基于1的索引）
# col2 = 2  # 第二列的索引（基于1的索引）
# delimiter = "_"  # 用于连接字符串的分隔符
# new_col = 3  # 合并后的新列的索引（基于1的索引）

# concatenate_columns(file_path, sheet_name, col1, col2, delimiter, new_col)

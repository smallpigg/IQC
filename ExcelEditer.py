from openpyxl import load_workbook

wb = load_workbook('list_V01.xlsx')
ws1 = wb['电子']
ws2 = wb['机械']

wb3 = load_workbook('TZD-AAA-A.xlsx')
ws3 = wb3["TB4.2.3-04"]

# 电子文件生成
for i in range(2,ws1.max_row+1):
    c = ws1.cell(row=i, column=4).value
    d = ws1.cell(row=i, column=10).value
    e = ws1.cell(row=i, column=13).value
    e = e.strftime("%Y-%m-%d")
    filename = "TB-"+d+"-A版 "+c+ "文件记录更改通知单.xlsx"
    c = "文件记录名称："+c+"进货检验作业指导书"
    d = "文件记录编号："+d
    e = "申请日期："+e

    ws3.cell(row=4, column=1).value = c
    ws3.cell(row=5, column=1).value = d
    ws3['F6'].value = e

    filename = "OUTPUT/"+filename
    wb3.save(filename)

# 机械文件生成
for i in range(2,ws2.max_row+1):
    c = ws2.cell(row=i, column=4).value
    d = ws2.cell(row=i, column=10).value
    e = ws2.cell(row=i, column=13).value
    e = e.strftime("%Y-%m-%d")
    filename = "TB-"+d+"-A版 "+c+ "文件记录更改通知单.xlsx"
    c = "文件记录名称："+c+"进货检验作业指导书"
    d = "文件记录编号："+d
    e = "申请日期："+e

    ws3.cell(row=4, column=1).value = c
    ws3.cell(row=5, column=1).value = d
    ws3['F6'].value = e

    filename = "OUTPUT_机械/"+filename
    wb3.save(filename)
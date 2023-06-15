import openpyxl

def replace_string_in_excel(file_path, sheet_name, find_str, replace_str):
    """
    将Excel工作表中的指定字符串替换为另一个字符串。

    :param file_path: Excel文件的路径
    :param sheet_name: 要操作的工作表名称
    :param find_str: 要查找的字符串
    :param replace_str: 用于替换的字符串
    """
    # 加载工作簿和工作表
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # 获取工作表的最大行和列数
    max_row = sheet.max_row
    max_col = sheet.max_column

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            # 获取单元格的值
            cell_value = sheet.cell(row=row, column=col).value

            # 检查单元格中是否包含要查找的字符串
            if cell_value and find_str in str(cell_value):
                # 替换字符串并将新值写回单元格
                new_value = cell_value.replace(find_str, replace_str)
                sheet.cell(row=row, column=col).value = new_value

    # 保存更改后的工作簿
    workbook.save(file_path)
    print("已完成字符串替换")

# # 示例用法
file_path = "result\\result-output.xlsx"  # 你的Excel文件路径
sheet_name = "Sheet1"  # 你要操作的工作表名称
find_str = "\n质量标准\nQuality Standard"  # 要查找的字符串
replace_str = ""  # 用于替换的字符串

replace_string_in_excel(file_path, sheet_name, find_str, replace_str)

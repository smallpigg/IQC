import openpyxl

def wrap_column_data(file_path, sheet_name, column, prefix, suffix, new_column):
    """
    在Excel工作表中指定列的数据前后添加指定字符，并将结果保存到新列。

    :param file_path: Excel文件的路径
    :param sheet_name: 要操作的工作表名称
    :param column: 要操作的列索引（基于1的索引）
    :param prefix: 要添加到数据前面的字符
    :param suffix: 要添加到数据后面的字符
    :param new_column: 结果保存的新列索引（基于1的索引）
    """
    # 加载工作簿和工作表
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # 获取工作表的最大行数
    max_row = sheet.max_row

    for row in range(1, max_row + 1):
        # 获取指定列的值
        cell_value = sheet.cell(row=row, column=column).value

        # 在值前后添加指定字符
        wrapped_value = prefix + str(cell_value) + suffix

        # 将包装后的值写入新列
        sheet.cell(row=row, column=new_column).value = wrapped_value

    # 保存更改后的工作簿
    workbook.save(file_path)
    print(f"处理了{max_row}行")

# 示例用法
file_path = "your_excel_file.xlsx"  # 你的Excel文件路径
sheet_name = "Sheet1"  # 你要操作的工作表名称
column = 1  # 要操作的列索引（基于1的索引）
prefix = "["  # 要添加到数据前面的字符
suffix = "]"  # 要添加到数据后面的字符
new_column = 2  # 结果保存的新列索引（基于1的索引）

wrap_column_data(file_path, sheet_name, column, prefix, suffix, new_column)
